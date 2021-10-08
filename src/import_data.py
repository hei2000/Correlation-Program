import openpyxl
import sys,json
import DataClasses
import logging
import helper
import datetime
import config

from PyQt5.QtWidgets import QApplication,QMainWindow,QWidget,QScrollArea,QGridLayout,QLabel,QComboBox,QCheckBox,QCompleter,QPushButton,QLineEdit,QMessageBox
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIntValidator

logging.basicConfig(filename="logging.log",level = logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')


class window(QMainWindow):
    def __init__(self,filepath):
        super().__init__()
        self.widget = QWidget()
        self.scroll = QScrollArea()
        self.grid = QGridLayout()
        
        self.FILEPATH = filepath
        
        self.initUI()
        
    def initUI(self):
        self.workbook = openpyxl.load_workbook(self.FILEPATH)
        self.sheet = self.workbook.active
        self.lab_id = self.sheet.cell(2,1).value
        self.lab_group = self.sheet.cell(2,2).value
        self.lab_location = self.sheet.cell(2,3).value
        self.lab_name = self.sheet.cell(2,4).value
        self.lab_third_party = self.sheet.cell(2,5).value
        self.TEMPLATE = [self.lab_id,self.lab_group,self.lab_location,self.lab_name,self.lab_third_party]
        self.REQUIREMENT_START = 4
        
        try:
            with open(config.BASE_DIRECTORY_PATH + 'Data/Tests.json') as json_file:  
                self.Tests = DataClasses.Tests.from_dict(json.load(json_file))
        except:
            logging.fatal("Cannot load data from Tests,closing 'Import Data'")
            exit()
    
        try:
            with open(config.BASE_DIRECTORY_PATH + 'Data/Labs.json') as json_file:  
                self.Labs = DataClasses.Labs.from_dict(json.load(json_file))
        except:
            logging.fatal("Cannot load data from ./Data/Labs.json,closing 'Import Data'")
            exit()
            
        bool_imported = self.check_imported()
        if bool_imported == -1:
            logging.fatal("Cannot find " + self.lab_id + " , closing 'Import Data'")
            exit()
        if bool_imported:
            logging.info("Imported " + self.lab_id + " already, closing 'Import Data'")
            message = QMessageBox()
            title = "Imported Already"
            message.setWindowTitle(title)
            message.setText("Cannot import the same lab more than once!")
            x = message.exec_()
            return -1
        max_row = self.sheet.max_row
        currentRow = 4
        while (currentRow <= max_row):
            template,sample,test_name,test_method = self.read_test(currentRow,self.sheet)
            self.write_test(template,sample,test_name,test_method)
            currentRow += 4
        self.record_imported()
            
    def check_imported(self):
        for lab in self.Labs.Groups[self.lab_group].values():
            if lab.ID == self.lab_id:
                return lab.imported
        return -1
    
    def read_test(self,currentRow,sheet):
        test_name = sheet.cell(currentRow + 1,1).value
        test_method = sheet.cell(currentRow + 1,2).value
        test = helper.find_test(self.Tests,test_name,test_method)
        sample = sheet.cell(currentRow + 1,3).value
        template = self.TEMPLATE.copy()
        template += [test_name,test_method,sample]
        requirement_column = self.REQUIREMENT_START
        for count in range(len(test.Requirements)):
            template.append(sheet.cell(currentRow + 2,requirement_column).value)
            requirement_column += 1
        result_column = requirement_column + 1
        for count in range(len(test.Results)):
            template.append(sheet.cell(currentRow + 2,result_column).value)
            result_column += 1
        template.append(sheet.cell(currentRow + 2,result_column).value)
    
        return template,sample,test_name,test_method
    
    def write_test(self,template,sample,test_name,test_method):
        filename = test_method + "_" + test_name + ".xlsx"
        filepath = config.BASE_DIRECTORY_PATH + "Data/Result/Sample " + sample + "/" + filename
        wb_obj = openpyxl.load_workbook(filepath)
        sheet_obj = wb_obj.active
        rows = sheet_obj.max_row
        row_count = rows + 1
        for column_count in range(1,len(template)+1):
            sheet_obj.cell(row_count,column_count).value = template[column_count-1]
            sheet_obj.cell(row_count,column_count).number_format = '@'
        wb_obj.save(filepath)
        
    def record_imported(self):
        try:
            for lab in self.Labs.Groups[self.lab_group].values():
                if lab.ID == self.lab_id:
                    lab.imported = True
                    #print(type(datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")))
                    lab.import_date = datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
                    break
            Labs_json = self.Labs.to_dict()
            with open(config.BASE_DIRECTORY_PATH + 'Data/Labs.json','w',newline='') as fp:
                json.dump(Labs_json,fp)
        except:
            logging.fatal("Cannot set lab.imported(" + self.lab_id + ") to be true")
            exit()
